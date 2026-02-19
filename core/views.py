from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.db.models import Q
from django.contrib import messages
from django.db import transaction
import json
from datetime import timedelta, datetime, date
import openpyxl 
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# --- IMPORTS MODEL & FORM ---
from django.contrib.auth.models import User, Group
from .models import Proyek, Tugas, TemplateBAU, AuditLog, UserProfile
from .forms import ProyekForm, TugasForm, ImportTugasForm, ImportUserForm

# --- HELPER & HIERARKI DIVISI ---
def log_activity(user, action, model_name, obj_id, details):
    AuditLog.objects.create(user=user, action=action, target_model=model_name, target_id=str(obj_id), details=details)

def get_role(user):
    try: return user.profile.role
    except Exception: return 'MEMBER'

def is_admin(user): return user.is_superuser or get_role(user) == 'ADMIN'
def is_leader(user): return get_role(user) == 'LEADER'
def is_member(user): return get_role(user) == 'MEMBER'

def get_accessible_groups(user):
    if user.is_superuser: return Group.objects.all()
    
    user_groups = list(user.groups.all())
    group_names = [g.name.upper() for g in user_groups]
    role = get_role(user)
    
    # Hierarki Risk Management
    if 'RISK MANAGEMENT' in group_names and role == 'ADMIN':
        sub_groups = Group.objects.filter(name__in=[
            'RISK PROCESS CONTROL', 'PORTFOLIO MANAGEMENT & GOVERNANCE', 'RISK PRODUCT & DEVELOPMENT'
        ])
        user_groups.extend(list(sub_groups))
        
    return Group.objects.filter(id__in=[g.id for g in user_groups]).distinct()

class GroupAccessMixin:
    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_superuser: return qs
        
        accessible_groups = get_accessible_groups(user)
        if self.model == Tugas:
            return qs.filter(Q(pemilik_grup__in=accessible_groups) | Q(ditugaskan_ke=user)).distinct()
        return qs.filter(pemilik_grup__in=accessible_groups).distinct()

@login_required
def dashboard(request):
    user = request.user
    if user.is_superuser:
        tasks = Tugas.objects.all()
        projects = Proyek.objects.all()
        team_members = User.objects.all()
    else:
        accessible_groups = get_accessible_groups(user)
        tasks = Tugas.objects.filter(Q(pemilik_grup__in=accessible_groups) | Q(ditugaskan_ke=user)).distinct()
        projects = Proyek.objects.filter(pemilik_grup__in=accessible_groups).distinct()
        team_members = User.objects.filter(groups__in=accessible_groups).distinct()
    
    assignee_id = request.GET.get('assignee')
    if assignee_id: tasks = tasks.filter(ditugaskan_ke_id=assignee_id)

    context = {
        'total_projects': projects.count(),
        'total_tasks': tasks.count(),
        'overdue_tasks': tasks.filter(status='OVERDUE').count(),
        'in_progress': tasks.filter(status='IN_PROGRESS').count(),
        'todo_count': tasks.filter(status='TODO').count(),
        'done_count': tasks.filter(status='DONE').count(),
        'on_hold_count': tasks.filter(status='ON_HOLD').count(), 
        'drop_count': tasks.filter(status='DROP').count(),
        'user_role': get_role(user),
        'team_members': team_members
    }
    return render(request, 'core/dashboard.html', context)

# --- PROYEK VIEWS ---
class ProyekListView(LoginRequiredMixin, GroupAccessMixin, ListView):
    model = Proyek
    template_name = 'core/proyek_list.html'
    context_object_name = 'proyek_list'

class ProyekCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Proyek
    form_class = ProyekForm
    template_name = 'core/proyek_form.html'
    success_url = reverse_lazy('proyek-list')
    
    def test_func(self): 
        # UPDATE: Member sekarang BOLEH membuat proyek (Return True)
        return True 
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        if not self.request.user.is_superuser:
            user_group = self.request.user.groups.first()
            if not user_group:
                form.add_error(None, "User tidak punya grup/divisi.")
                return self.form_invalid(form)
            form.instance.pemilik_grup = user_group
            
        form.instance.dibuat_oleh = self.request.user
        resp = super().form_valid(form)
        log_activity(self.request.user, 'CREATE', 'Proyek', self.object.kode_proyek, f"Created: {self.object.nama_proyek}")
        return resp

class ProyekUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Proyek
    form_class = ProyekForm
    template_name = 'core/proyek_form.html'
    success_url = reverse_lazy('proyek-list')
    
    def test_func(self): 
        user = self.request.user
        # UPDATE: Admin/Leader/Superuser akses penuh. Member hanya jika dia PEMBUATNYA.
        if user.is_superuser or is_admin(user) or is_leader(user): return True
        return self.get_object().dibuat_oleh == user
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
        
    def form_valid(self, form):
        resp = super().form_valid(form)
        log_activity(self.request.user, 'UPDATE', 'Proyek', self.object.kode_proyek, "Updated details")
        return resp

class ProyekDetailView(LoginRequiredMixin, GroupAccessMixin, DetailView):
    model = Proyek
    template_name = 'core/proyek_detail.html'

class ProyekDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Proyek
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('proyek-list')
    
    def test_func(self): 
        # UPDATE: Hanya Leader ke atas yang boleh HAPUS
        return is_admin(self.request.user) or is_leader(self.request.user) or self.request.user.is_superuser
        
    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        log_activity(request.user, 'DELETE', 'Proyek', obj.kode_proyek, f"Deleted: {obj.nama_proyek}")
        return super().delete(request, *args, **kwargs)

# --- IMPORT TUGAS VIEWS (UPDATED: MEMBER ACCESS & SUBTASK) ---
@login_required
def download_template_tugas(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Template_Import_Tugas_Subtask.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Tugas"
    
    # UPDATE: Tambahkan kolom Level dan Nama Induk
    headers = [
        'Nama Tugas', 'Tipe Tugas (PROJECT/ADHOC/BAU)', 'Kode Proyek', 
        'Pemberi Tugas', 'Username PIC', 'Start Date', 'End Date', 'Deskripsi',
        'Level (1=Main, 2=Sub)', 'Nama Tugas Induk (Wajib jika Level 2)'
    ]
    ws.append(headers)
    
    # Styling Header
    for cell in ws[1]: 
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Sample Data
    sample_data = [
        ['Setup Server Utama', 'PROJECT', 'P-001', '', request.user.username, '2025-02-10', '2025-02-14', 'Ini tugas induk', 1, ''],
        ['Install Database', 'PROJECT', 'P-001', '', request.user.username, '2025-02-11', '2025-02-12', 'Ini subtask', 2, 'Setup Server Utama'],
        ['Laporan Mingguan', 'ADHOC', '', 'Pak Boss', '', '2025-02-10', '2025-02-10', 'Tugas biasa', 1, ''],
    ]
    for row in sample_data: ws.append(row)
    
    # Auto Width
    for column in ws.columns:
        ws.column_dimensions[get_column_letter(column[0].column)].width = 25
        
    wb.save(response)
    return response

@login_required
def import_tugas(request):
    # UPDATE: Member sekarang BOLEH import tugas
    if not (is_admin(request.user) or is_leader(request.user) or is_member(request.user) or request.user.is_superuser): 
        return HttpResponseForbidden("Akses ditolak.")

    if request.method == 'POST':
        form = ImportTugasForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = openpyxl.load_workbook(request.FILES['file_excel'])
                ws = wb.active
                
                # UPDATE: Logika Sorting (Level 1 dulu, baru Level 2)
                # Kita baca dulu semua baris ke dalam list of dictionary
                raw_rows = []
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]: continue
                    # Mapping kolom
                    item = {
                        'idx': idx,
                        'nama': row[0],
                        'tipe': (row[1] or 'ADHOC').upper().strip(),
                        'kode_p': row[2],
                        'pemberi': row[3],
                        'pic_uname': row[4],
                        'start': row[5],
                        'end': row[6],
                        'desc': row[7] or "",
                        # Kolom baru (index 8 dan 9)
                        'level': int(row[8]) if len(row) > 8 and row[8] else 1,
                        'parent_name': str(row[9]).strip() if len(row) > 9 and row[9] else None
                    }
                    raw_rows.append(item)
                
                # SORTING: Proses Level 1 (Main Task) duluan agar parent tersedia saat Subtask dibuat
                raw_rows.sort(key=lambda x: x['level'])
                
                success_count = 0
                errors = []
                created_tasks_cache = {} # Cache untuk menyimpan tugas yg baru dibuat di batch ini
                
                user_group = request.user.groups.first()
                if not user_group and not request.user.is_superuser:
                    raise ValueError("User Anda tidak terdaftar dalam Divisi/Group manapun.")

                with transaction.atomic():
                    for row in raw_rows:
                        try:
                            # 1. Parsing Tanggal
                            def parse_date(d):
                                if isinstance(d, datetime): return d.date()
                                if isinstance(d, str):
                                    for f in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'): 
                                        try: return datetime.strptime(d.strip(), f).date()
                                        except: pass
                                return d if isinstance(d, date) else None

                            start = parse_date(row['start'])
                            end = parse_date(row['end'])
                            
                            if not start or not end: raise ValueError("Format tanggal salah")
                            if start.weekday() >= 5: raise ValueError("Tanggal Mulai hari libur (Sabtu/Minggu)")

                            # 2. Cek Proyek
                            proyek_obj = None
                            if row['tipe'] == 'PROJECT':
                                if not row['kode_p']: raise ValueError("Kode Proyek wajib diisi utk tipe PROJECT")
                                proyek_obj = Proyek.objects.filter(kode_proyek=row['kode_p']).first()
                                if not proyek_obj: raise ValueError(f"Proyek {row['kode_p']} tidak ditemukan")
                            
                            # 3. Cek PIC
                            pic_user = None
                            if row['pic_uname']:
                                pic_user = User.objects.filter(username=row['pic_uname']).first()
                                
                            # 4. Handle Subtask (Level 2)
                            parent_obj = None
                            if row['level'] == 2:
                                if not row['parent_name']:
                                    raise ValueError("Level 2 (Subtask) wajib mengisi 'Nama Tugas Induk'")
                                
                                # Cari di cache batch ini dulu
                                if row['parent_name'] in created_tasks_cache:
                                    parent_obj = created_tasks_cache[row['parent_name']]
                                else:
                                    # Cari di database (di grup yg sama)
                                    # Gunakan filter pemilik_grup agar tidak cross-division
                                    qs = Tugas.objects.filter(nama_tugas=row['parent_name'])
                                    if not request.user.is_superuser:
                                        qs = qs.filter(pemilik_grup=user_group)
                                    
                                    parent_obj = qs.first()
                                    
                                if not parent_obj:
                                    raise ValueError(f"Tugas Induk '{row['parent_name']}' tidak ditemukan.")

                                # Warisi Proyek dari Induk jika tidak diisi
                                if not proyek_obj and parent_obj.proyek:
                                    proyek_obj = parent_obj.proyek

                            # 5. Create
                            new_task = Tugas.objects.create(
                                nama_tugas=row['nama'],
                                tipe_tugas=row['tipe'],
                                proyek=proyek_obj,
                                induk=parent_obj, # Link ke Parent
                                pemberi_tugas=row['pemberi'] or request.user.get_full_name(),
                                ditugaskan_ke=pic_user,
                                tanggal_mulai=start,
                                tenggat_waktu=end,
                                deskripsi=row['desc'] if hasattr(Tugas, 'deskripsi') else "", # Optional check
                                pemilik_grup=user_group if user_group else (proyek_obj.pemilik_grup if proyek_obj else None),
                                status='TODO',
                                progress=0
                            )
                            
                            # Simpan ke cache agar bisa jadi induk bagi baris berikutnya
                            created_tasks_cache[row['nama']] = new_task
                            success_count += 1
                            
                        except Exception as e:
                            errors.append(f"Baris {row['idx']} ({row['nama']}): {str(e)}")
                
                if success_count: messages.success(request, f"Sukses import {success_count} tugas.")
                if errors: messages.warning(request, f"Gagal {len(errors)} data: " + "; ".join(errors[:3]))
                return redirect('tugas-list')

            except Exception as e: messages.error(request, f"File Error: {str(e)}")
    else: form = ImportTugasForm()
    return render(request, 'core/import_tugas.html', {'form': form})

# --- USER IMPORT & MANAGEMENT ---
@login_required
def download_template_user(request):
    if not (request.user.is_superuser or is_admin(request.user)): return HttpResponseForbidden("Akses ditolak.")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Template_Import_User.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Template User"
    ws.append(['Username', 'Email', 'Password', 'First Name', 'Last Name', 'Role (ADMIN/LEADER/MEMBER)', 'Nama Divisi (Group)', 'Status (ACTIVE/INACTIVE)'])
    for cell in ws[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    wb.save(response)
    return response

@login_required
def import_user(request):
    if not request.user.is_superuser: return HttpResponseForbidden("Hanya Superuser.")
    if request.method == 'POST':
        form = ImportUserForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = openpyxl.load_workbook(request.FILES['file_excel'])
                ws = wb.active
                success_users, errors = [], []
                
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]: continue
                    try:
                        with transaction.atomic():
                            uname = str(row[0]).strip().lower().replace(" ", "")
                            email, pwd = row[1], str(row[2]) if row[2] else "Default123"
                            fname, lname = row[3] or "", row[4] or ""
                            role, group_name = (str(row[5]).upper().strip() if row[5] else 'MEMBER'), str(row[6]).strip() if row[6] else None
                            
                            status_staf = str(row[7]).upper().strip() if len(row) > 7 and row[7] else 'ACTIVE'
                            is_active_user = False if status_staf in ['INACTIVE', 'NONAKTIF', '0', 'FALSE'] else True

                            if User.objects.filter(username=uname).exists(): raise ValueError(f"Username {uname} sudah ada")
                            
                            u = User.objects.create_user(username=uname, email=email, password=pwd)
                            u.first_name = fname; u.last_name = lname
                            u.is_active = is_active_user 
                            u.is_staff = True 
                            u.save()
                            
                            p, created = UserProfile.objects.get_or_create(user=u)
                            p.role = role; p.save()
                            
                            if group_name:
                                g, _ = Group.objects.get_or_create(name=group_name)
                                u.groups.add(g)
                            
                            success_users.append(uname)
                    except Exception as e: errors.append(f"Baris {idx} ({row[0]}): {str(e)}")
                
                if success_users: messages.success(request, f"Sukses buat user: {', '.join(success_users[:5])}...")
                if errors: messages.warning(request, f"Gagal: {'; '.join(errors[:5])}")
                return redirect('user-list')

            except Exception as e: messages.error(request, f"File Error: {str(e)}")
    else: form = ImportUserForm()
    return render(request, 'core/import_user.html', {'form': form})

class UserListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = User
    template_name = 'core/user_list.html'
    context_object_name = 'users'
    def test_func(self): return self.request.user.is_superuser
    def get_queryset(self): return User.objects.all().order_by('username').select_related('profile')

@login_required
def bulk_delete_users(request):
    if not request.user.is_superuser: return HttpResponseForbidden("Akses ditolak.")
    if request.method == 'POST':
        user_ids = request.POST.getlist('selected_users')
        if user_ids:
            users_to_delete = User.objects.filter(id__in=user_ids).exclude(id=request.user.id)
            count = users_to_delete.count()
            if count > 0:
                users_to_delete.delete()
                messages.success(request, f"Berhasil menghapus {count} pengguna.")
            else: messages.warning(request, "Tidak ada data yang dihapus.")
        else: messages.warning(request, "Tidak ada pengguna yang dipilih.")
    return redirect('user-list')

# --- TUGAS VIEWS ---
class TugasListView(LoginRequiredMixin, GroupAccessMixin, ListView):
    model = Tugas
    template_name = 'core/tugas_list.html'
    context_object_name = 'tugas_list'
    
    def get_queryset(self):
        qs = super().get_queryset()
        assignee_id = self.request.GET.get('assignee')
        if assignee_id: qs = qs.filter(ditugaskan_ke_id=assignee_id)
        return qs.order_by('kode_tugas')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_superuser:
            context['team_members'] = User.objects.filter(is_active=True).order_by('first_name')
        else:
            context['team_members'] = User.objects.filter(groups__in=get_accessible_groups(self.request.user), is_active=True).distinct()
        return context

class TugasCreateView(LoginRequiredMixin, CreateView):
    model = Tugas
    form_class = TugasForm
    template_name = 'core/tugas_form.html'
    success_url = reverse_lazy('tugas-list')
    
    def get_initial(self):
        initial = super().get_initial()
        initial['pemberi_tugas'] = self.request.user.get_full_name() or self.request.user.username
        parent_id = self.request.GET.get('parent_id')
        if parent_id:
            try:
                parent = Tugas.objects.get(pk=parent_id)
                initial['induk'], initial['proyek'], initial['tanggal_mulai'] = parent, parent.proyek, parent.tanggal_mulai
            except: pass
        return initial
        
    def get_form_kwargs(self): 
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
        
    def form_valid(self, form):
        if not self.request.user.is_superuser:
            user_group = self.request.user.groups.first()
            if not user_group: 
                form.add_error(None, "User tidak punya grup.")
                return self.form_invalid(form)
            form.instance.pemilik_grup = user_group
            
        log_activity(self.request.user, 'CREATE', 'Tugas', form.instance.kode_tugas, f"Created: {form.instance.nama_tugas}")
        return super().form_valid(form)

class TugasUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Tugas
    form_class = TugasForm
    template_name = 'core/tugas_form.html'
    success_url = reverse_lazy('tugas-list')
    
    def get_form_kwargs(self): 
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
        
    def test_func(self): 
        user = self.request.user
        if user.is_superuser or is_admin(user) or is_leader(user): return True
        # Member boleh edit tugas sendiri ATAU tugas yang masih kosong (unassigned)
        return self.get_object().ditugaskan_ke == user or self.get_object().ditugaskan_ke is None
        
    def dispatch(self, request, *args, **kwargs):
        if self.get_object().status == 'DONE' and not request.user.is_superuser: 
            messages.warning(request, "Tugas SELESAI tidak bisa diedit.")
            return redirect('tugas-list')
        return super().dispatch(request, *args, **kwargs)
        
    def form_valid(self, form):
        if form.has_changed(): 
            log_activity(self.request.user, 'UPDATE', 'Tugas', self.object.kode_tugas, f"Changed: {', '.join(form.changed_data)}")
        return super().form_valid(form)

class TugasDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Tugas
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('tugas-list')
    
    def test_func(self): 
        return is_admin(self.request.user) or self.request.user.is_superuser
        
    def delete(self, request, *args, **kwargs):
        log_activity(request.user, 'DELETE', 'Tugas', self.get_object().kode_tugas, f"Deleted: {self.get_object().nama_tugas}")
        return super().delete(request, *args, **kwargs)

# --- API HELPERS ---
@login_required
def get_entity_dates_api(request): return JsonResponse({}) 

@login_required
def update_progress_api(request, pk):
    if request.method == 'POST':
        try:
            d = json.loads(request.body); prog = int(d.get('progress', 0))
            t = get_object_or_404(Tugas, pk=pk)
            t.progress = prog
            if prog == 100: t.status = 'DONE'
            elif prog > 0 and t.status == 'TODO': t.status = 'IN_PROGRESS'
            elif prog == 0: t.status = 'TODO'
            t.save()
            log_activity(request.user, 'UPDATE', 'Tugas', t.kode_tugas, f"Progress: {prog}%")
            return JsonResponse({'status': 'success', 'new_status': t.get_status_display()})
        except Exception as e: return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid'}, status=405)

@login_required
def update_task_date_api(request, pk):
    if request.method == 'POST':
        try:
            d = json.loads(request.body)
            s = datetime.strptime(d.get('start'), "%Y-%m-%d").date()
            e = datetime.strptime(d.get('end'), "%Y-%m-%d").date()
            if s.weekday() >= 5: return JsonResponse({'error': 'Hari Libur!'}, status=400)
            t = get_object_or_404(Tugas, pk=pk)
            if not (request.user.is_superuser or is_admin(request.user) or is_leader(request.user) or t.ditugaskan_ke == request.user):
                return JsonResponse({'error': 'Permission denied'}, status=403)
            t.tanggal_mulai = s; t.tenggat_waktu = e; t.save()
            log_activity(request.user, 'UPDATE', 'Tugas', t.kode_tugas, f"Gantt: {s}->{e}")
            return JsonResponse({'status': 'success'})
        except Exception as e: return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid'}, status=405)

# --- GANTT & CALENDAR ---
@login_required
def gantt_data(request):
    user = request.user
    if user.is_superuser: tasks = Tugas.objects.all()
    else:
        accessible_groups = get_accessible_groups(user)
        tasks = Tugas.objects.filter(Q(pemilik_grup__in=accessible_groups) | Q(ditugaskan_ke=user)).distinct()
        
    if request.GET.get('assignee'): tasks = tasks.filter(ditugaskan_ke_id=request.GET.get('assignee'))
    
    gantt_list = []
    visible_ids = set(tasks.values_list('id', flat=True))
    
    for p in Proyek.objects.filter(id__in=tasks.values_list('proyek_id', flat=True)).distinct():
        gantt_list.append({'id': f"P-{p.id}", 'name': f"📁 {p.nama_proyek}", 'start': str(p.tanggal_mulai), 'end': str(p.tanggal_selesai), 'progress': 0, 'custom_class': 'bar-project', 'read_only': True})
        for t in tasks.filter(proyek=p):
            dep = str(t.tergantung_pada.id) if t.tergantung_pada and t.tergantung_pada.id in visible_ids else ""
            cls = 'bar-done' if t.status=='DONE' else ('bar-overdue' if t.status=='OVERDUE' else ('bar-hold' if t.status=='ON_HOLD' else ''))
            gantt_list.append({'id': str(t.id), 'name': t.nama_tugas, 'start': str(t.tanggal_mulai), 'end': str(t.tenggat_waktu), 'progress': t.progress, 'dependencies': dep, 'custom_class': cls})
            
    for t in tasks.filter(proyek__isnull=True):
        dep = str(t.tergantung_pada.id) if t.tergantung_pada and t.tergantung_pada.id in visible_ids else ""
        cls = 'bar-done' if t.status=='DONE' else ('bar-overdue' if t.status=='OVERDUE' else ('bar-hold' if t.status=='ON_HOLD' else ''))
        gantt_list.append({'id': str(t.id), 'name': t.nama_tugas, 'start': str(t.tanggal_mulai), 'end': str(t.tenggat_waktu), 'progress': t.progress, 'dependencies': dep, 'custom_class': cls})

    return JsonResponse(gantt_list, safe=False)

@login_required
def gantt_view(request): 
    if request.user.is_superuser: team = User.objects.all()
    else: team = User.objects.filter(groups__in=get_accessible_groups(request.user)).distinct()
    return render(request, 'core/gantt.html', {'team_members': team})

@login_required
def export_gantt_excel(request): return HttpResponse("Export OK")

@login_required
def calendar_view(request): return render(request, 'core/calendar.html')

@login_required
def calendar_data(request):
    user = request.user
    if user.is_superuser: tasks = Tugas.objects.all()
    else: tasks = Tugas.objects.filter(Q(pemilik_grup__in=get_accessible_groups(user)) | Q(ditugaskan_ke=user)).distinct()
    
    events = []
    for t in tasks:
        color = '#0d6efd' 
        if t.status == 'DONE': color = '#198754' 
        elif t.status == 'OVERDUE': color = '#dc3545' 
        elif t.status == 'ON_HOLD': color = '#ffc107' 
        elif t.status == 'IN_PROGRESS': color = '#0dcaf0' 
        
        end_date = t.tenggat_waktu + timedelta(days=1)
        events.append({
            'title': f"{t.nama_tugas} ({t.progress}%)",
            'start': str(t.tanggal_mulai),
            'end': str(end_date), 
            'backgroundColor': color,
            'borderColor': color,
            'url': f"/tugas/{t.pk}/update/" 
        })
    return JsonResponse(events, safe=False)

# --- BAU VIEWS ---
class TemplateBAUListView(LoginRequiredMixin, GroupAccessMixin, ListView):
    model = TemplateBAU
    template_name = 'core/bau_list.html'
    context_object_name = 'bau_list'

class TemplateBAUCreateView(LoginRequiredMixin, CreateView):
    model = TemplateBAU
    fields = ['nama_tugas', 'deskripsi', 'frekuensi', 'default_pic']
    template_name = 'core/bau_form.html'
    success_url = reverse_lazy('bau-list')
    def form_valid(self, form):
        form.instance.pemilik_grup = self.request.user.groups.first()
        return super().form_valid(form)

class TemplateBAUUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = TemplateBAU
    fields = ['nama_tugas', 'deskripsi', 'frekuensi', 'default_pic']
    template_name = 'core/bau_form.html'
    success_url = reverse_lazy('bau-list')
    def test_func(self): return is_admin(self.request.user)

class TemplateBAUDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = TemplateBAU
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('bau-list')
    def test_func(self): return is_admin(self.request.user)

@login_required
def trigger_bau_single(request, pk):
    return redirect('bau-list')